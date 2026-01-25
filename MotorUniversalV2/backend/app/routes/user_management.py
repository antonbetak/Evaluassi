"""
Rutas para gestión de usuarios (admin y coordinadores)
"""
from flask import Blueprint, request, jsonify, g
from functools import wraps
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models import User
import uuid
import re

bp = Blueprint('user_management', __name__, url_prefix='/api/user-management')

# Roles disponibles en el sistema (sin alumno)
AVAILABLE_ROLES = ['admin', 'editor', 'soporte', 'coordinator', 'candidato', 'auxiliar']

# Roles que puede crear cada tipo de usuario
ROLE_CREATE_PERMISSIONS = {
    'admin': ['editor', 'soporte', 'coordinator', 'candidato', 'auxiliar'],  # Todo menos admin
    'coordinator': ['candidato']  # Solo candidatos
}


def management_required(f):
    """Decorador que requiere rol de admin o coordinator"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'No autorizado'}), 401
        if user.role not in ['admin', 'coordinator']:
            return jsonify({'error': 'Acceso denegado. Se requiere rol de administrador o coordinador'}), 403
        g.current_user = user
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Decorador que requiere rol de admin"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'No autorizado'}), 401
        if user.role != 'admin':
            return jsonify({'error': 'Acceso denegado. Se requiere rol de administrador'}), 403
        g.current_user = user
        return f(*args, **kwargs)
    return decorated


def validate_email(email):
    """Validar formato de email"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_password(password):
    """Validar requisitos de contraseña"""
    if len(password) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres"
    if not re.search(r'[A-Z]', password):
        return False, "La contraseña debe tener al menos una mayúscula"
    if not re.search(r'[a-z]', password):
        return False, "La contraseña debe tener al menos una minúscula"
    if not re.search(r'[0-9]', password):
        return False, "La contraseña debe tener al menos un número"
    return True, None


# ============== LISTAR USUARIOS ==============

@bp.route('/users', methods=['GET'])
@jwt_required()
@management_required
def list_users():
    """Listar usuarios según permisos del solicitante"""
    try:
        current_user = g.current_user
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        search = request.args.get('search', '')
        role_filter = request.args.get('role', '')
        active_filter = request.args.get('is_active', '')
        
        query = User.query
        
        # Coordinadores solo ven candidatos
        if current_user.role == 'coordinator':
            query = query.filter(User.role == 'candidato')
        
        # Filtros
        if role_filter:
            # Soportar múltiples roles separados por coma (ej: "admin,editor,soporte")
            roles = [r.strip() for r in role_filter.split(',') if r.strip()]
            if len(roles) == 1:
                query = query.filter(User.role == roles[0])
            elif len(roles) > 1:
                query = query.filter(User.role.in_(roles))
            
        if active_filter != '':
            is_active = active_filter.lower() == 'true'
            query = query.filter(User.is_active == is_active)
        
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                db.or_(
                    User.name.ilike(search_term),
                    User.first_surname.ilike(search_term),
                    User.email.ilike(search_term),
                    User.curp.ilike(search_term),
                    User.username.ilike(search_term)
                )
            )
        
        query = query.order_by(User.created_at.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'users': [u.to_dict(include_private=True) for u in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/users/<string:user_id>', methods=['GET'])
@jwt_required()
@management_required
def get_user_detail(user_id):
    """Obtener detalle de un usuario"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        
        # Coordinadores solo pueden ver candidatos
        if current_user.role == 'coordinator' and user.role != 'candidato':
            return jsonify({'error': 'No tienes permiso para ver este usuario'}), 403
        
        return jsonify({
            'user': user.to_dict(include_private=True, include_partners=True)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== CREAR USUARIOS ==============

@bp.route('/users', methods=['POST'])
@jwt_required()
@management_required
def create_user():
    """Crear un nuevo usuario"""
    try:
        current_user = g.current_user
        data = request.get_json()
        
        # Validar campos requeridos
        required_fields = ['email', 'password', 'name', 'first_surname', 'role']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'El campo {field} es requerido'}), 400
        
        email = data['email'].strip().lower()
        role = data['role']
        
        # Validar email
        if not validate_email(email):
            return jsonify({'error': 'Formato de email inválido'}), 400
        
        # Verificar email único
        if User.query.filter_by(email=email).first():
            return jsonify({'error': 'Ya existe un usuario con ese email'}), 400
        
        # Validar contraseña
        is_valid, error_msg = validate_password(data['password'])
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        # Verificar permisos de creación según rol
        allowed_roles = ROLE_CREATE_PERMISSIONS.get(current_user.role, [])
        if role not in allowed_roles:
            if current_user.role == 'coordinator':
                return jsonify({'error': 'Los coordinadores solo pueden crear usuarios de tipo candidato'}), 403
            else:
                return jsonify({'error': f'No tienes permiso para crear usuarios con rol {role}'}), 403
        
        # Verificar CURP único si se proporciona
        if data.get('curp'):
            curp = data['curp'].upper().strip()
            if User.query.filter_by(curp=curp).first():
                return jsonify({'error': 'Ya existe un usuario con ese CURP'}), 400
        
        # Generar username si no se proporciona
        username = data.get('username', '').strip()
        if not username:
            # Generar username a partir del email
            base_username = email.split('@')[0]
            username = base_username
            counter = 1
            while User.query.filter_by(username=username).first():
                username = f"{base_username}{counter}"
                counter += 1
        else:
            # Verificar username único
            if User.query.filter_by(username=username).first():
                return jsonify({'error': 'Ya existe un usuario con ese nombre de usuario'}), 400
        
        # Crear usuario
        new_user = User(
            id=str(uuid.uuid4()),
            email=email,
            username=username,
            name=data['name'].strip(),
            first_surname=data['first_surname'].strip(),
            second_surname=data.get('second_surname', '').strip() or None,
            gender=data.get('gender'),
            curp=data.get('curp', '').upper().strip() or None,
            phone=data.get('phone', '').strip() or None,
            role=role,
            is_active=data.get('is_active', True),
            is_verified=data.get('is_verified', False)
        )
        new_user.set_password(data['password'])
        
        db.session.add(new_user)
        db.session.commit()
        
        return jsonify({
            'message': 'Usuario creado exitosamente',
            'user': new_user.to_dict(include_private=True)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== ACTUALIZAR USUARIOS ==============

@bp.route('/users/<string:user_id>', methods=['PUT'])
@jwt_required()
@management_required
def update_user(user_id):
    """Actualizar un usuario"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        
        # Coordinadores solo pueden editar candidatos
        if current_user.role == 'coordinator' and user.role != 'candidato':
            return jsonify({'error': 'No tienes permiso para editar este usuario'}), 403
        
        # No se puede editar a uno mismo por esta ruta (usar perfil)
        if user_id == current_user.id:
            return jsonify({'error': 'Usa la opción de perfil para editar tu propia cuenta'}), 400
        
        # Campos actualizables
        basic_fields = ['name', 'first_surname', 'second_surname', 'gender', 'phone']
        for field in basic_fields:
            if field in data:
                setattr(user, field, data[field].strip() if data[field] else None)
        
        # CURP - verificar unicidad
        if 'curp' in data:
            curp = data['curp'].upper().strip() if data['curp'] else None
            if curp:
                existing = User.query.filter(User.curp == curp, User.id != user_id).first()
                if existing:
                    return jsonify({'error': 'Ya existe un usuario con ese CURP'}), 400
            user.curp = curp
        
        # Email - verificar unicidad
        if 'email' in data:
            email = data['email'].strip().lower()
            if not validate_email(email):
                return jsonify({'error': 'Formato de email inválido'}), 400
            existing = User.query.filter(User.email == email, User.id != user_id).first()
            if existing:
                return jsonify({'error': 'Ya existe un usuario con ese email'}), 400
            user.email = email
        
        # Campos que solo admin puede cambiar
        if current_user.role == 'admin':
            if 'role' in data:
                new_role = data['role']
                # Admin no puede crear otros admins
                if new_role == 'admin' and user.role != 'admin':
                    return jsonify({'error': 'No se puede asignar rol de administrador'}), 403
                user.role = new_role
            
            if 'is_active' in data:
                user.is_active = data['is_active']
            
            if 'is_verified' in data:
                user.is_verified = data['is_verified']
        
        # Coordinadores solo pueden cambiar is_active de candidatos
        elif current_user.role == 'coordinator':
            if 'is_active' in data:
                user.is_active = data['is_active']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Usuario actualizado exitosamente',
            'user': user.to_dict(include_private=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== CAMBIAR CONTRASEÑA ==============

@bp.route('/users/<string:user_id>/password', methods=['PUT'])
@jwt_required()
@management_required
def change_user_password(user_id):
    """Cambiar contraseña de un usuario (sin requerir contraseña actual)"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        
        # Coordinadores solo pueden cambiar contraseña de candidatos
        if current_user.role == 'coordinator' and user.role != 'candidato':
            return jsonify({'error': 'No tienes permiso para cambiar la contraseña de este usuario'}), 403
        
        new_password = data.get('new_password')
        if not new_password:
            return jsonify({'error': 'La nueva contraseña es requerida'}), 400
        
        is_valid, error_msg = validate_password(new_password)
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        user.set_password(new_password)
        db.session.commit()
        
        return jsonify({
            'message': 'Contraseña actualizada exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== ACTIVAR/DESACTIVAR USUARIOS ==============

@bp.route('/users/<string:user_id>/toggle-active', methods=['POST'])
@jwt_required()
@management_required
def toggle_user_active(user_id):
    """Activar o desactivar un usuario"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        
        # Coordinadores solo pueden manejar candidatos
        if current_user.role == 'coordinator' and user.role != 'candidato':
            return jsonify({'error': 'No tienes permiso para modificar este usuario'}), 403
        
        # No se puede desactivar a uno mismo
        if user_id == current_user.id:
            return jsonify({'error': 'No puedes desactivarte a ti mismo'}), 400
        
        user.is_active = not user.is_active
        db.session.commit()
        
        status = 'activado' if user.is_active else 'desactivado'
        
        return jsonify({
            'message': f'Usuario {status} exitosamente',
            'user': user.to_dict(include_private=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== OPCIONES DE DOCUMENTOS ==============

@bp.route('/users/<string:user_id>/document-options', methods=['PUT'])
@jwt_required()
@admin_required
def update_user_document_options(user_id):
    """Actualizar opciones de documentos de un usuario (solo admin)"""
    try:
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        
        if 'evaluation_report' in data:
            user.enable_evaluation_report = data['evaluation_report']
        if 'certificate' in data:
            user.enable_certificate = data['certificate']
        if 'conocer_certificate' in data:
            user.enable_conocer_certificate = data['conocer_certificate']
        if 'digital_badge' in data:
            user.enable_digital_badge = data['digital_badge']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Opciones de documentos actualizadas',
            'user': user.to_dict(include_private=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== ESTADÍSTICAS ==============

@bp.route('/stats', methods=['GET'])
@jwt_required()
@management_required
def get_user_stats():
    """Obtener estadísticas de usuarios"""
    try:
        current_user = g.current_user
        
        base_query = User.query
        
        # Coordinadores solo ven stats de candidatos
        if current_user.role == 'coordinator':
            base_query = base_query.filter(User.role == 'candidato')
        
        total_users = base_query.count()
        active_users = base_query.filter(User.is_active == True).count()
        inactive_users = base_query.filter(User.is_active == False).count()
        verified_users = base_query.filter(User.is_verified == True).count()
        
        # Usuarios por rol (solo para admin)
        users_by_role = []
        if current_user.role == 'admin':
            for role in AVAILABLE_ROLES:
                count = User.query.filter_by(role=role).count()
                users_by_role.append({'role': role, 'count': count})
        else:
            users_by_role = [{'role': 'candidato', 'count': total_users}]
        
        return jsonify({
            'total_users': total_users,
            'active_users': active_users,
            'inactive_users': inactive_users,
            'verified_users': verified_users,
            'users_by_role': users_by_role
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== ROLES DISPONIBLES ==============

@bp.route('/roles', methods=['GET'])
@jwt_required()
@management_required
def get_available_roles():
    """Obtener roles que puede crear el usuario actual"""
    try:
        current_user = g.current_user
        allowed_roles = ROLE_CREATE_PERMISSIONS.get(current_user.role, [])
        
        role_descriptions = {
            'admin': 'Administrador - Acceso total al sistema',
            'editor': 'Editor - Gestión de exámenes y contenidos',
            'soporte': 'Soporte - Atención a usuarios y vouchers',
            'coordinator': 'Coordinador - Gestión de partners y candidatos',
            'candidato': 'Candidato - Usuario que presenta evaluaciones',
            'auxiliar': 'Auxiliar - Acceso de solo lectura'
        }
        
        return jsonify({
            'roles': [
                {'value': role, 'label': role.capitalize(), 'description': role_descriptions.get(role, '')}
                for role in allowed_roles
            ],
            'all_roles': [
                {'value': role, 'label': role.capitalize(), 'description': role_descriptions.get(role, '')}
                for role in AVAILABLE_ROLES
            ] if current_user.role == 'admin' else None
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== CARGA MASIVA DE CANDIDATOS ==============

@bp.route('/candidates/bulk-upload', methods=['POST'])
@jwt_required()
@management_required
def bulk_upload_candidates():
    """
    Carga masiva de candidatos desde archivo Excel
    
    Formato esperado del Excel (columnas):
    - email (requerido)
    - nombre (requerido)  
    - primer_apellido (requerido)
    - segundo_apellido (opcional)
    - genero (opcional: M, F, O)
    - curp (opcional)
    - telefono (opcional)
    - password (opcional, si no se proporciona se genera uno automático)
    """
    try:
        import io
        import secrets
        import string
        from openpyxl import load_workbook
        
        current_user = g.current_user
        
        # Verificar que se envió un archivo
        if 'file' not in request.files:
            return jsonify({'error': 'No se envió ningún archivo'}), 400
        
        file = request.files['file']
        
        if not file.filename:
            return jsonify({'error': 'No se seleccionó ningún archivo'}), 400
        
        # Verificar extensión
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'El archivo debe ser formato Excel (.xlsx o .xls)'}), 400
        
        # Leer el archivo Excel
        try:
            workbook = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
            sheet = workbook.active
        except Exception as e:
            return jsonify({'error': f'Error al leer el archivo Excel: {str(e)}'}), 400
        
        # Obtener encabezados (primera fila)
        headers = []
        for cell in sheet[1]:
            value = cell.value
            if value:
                headers.append(str(value).strip().lower())
            else:
                headers.append('')
        
        # Mapeo de nombres de columna esperados
        column_mapping = {
            'email': ['email', 'correo', 'correo electronico', 'correo electrónico', 'e-mail'],
            'nombre': ['nombre', 'name', 'nombres'],
            'primer_apellido': ['primer_apellido', 'primer apellido', 'apellido paterno', 'apellido_paterno', 'first_surname', 'apellido1'],
            'segundo_apellido': ['segundo_apellido', 'segundo apellido', 'apellido materno', 'apellido_materno', 'second_surname', 'apellido2'],
            'genero': ['genero', 'género', 'sexo', 'gender'],
            'curp': ['curp'],
            'telefono': ['telefono', 'teléfono', 'phone', 'celular', 'cel'],
            'password': ['password', 'contraseña', 'clave']
        }
        
        # Encontrar índices de columnas
        column_indices = {}
        for field, aliases in column_mapping.items():
            for i, header in enumerate(headers):
                if header in aliases:
                    column_indices[field] = i
                    break
        
        # Verificar columnas requeridas
        required_columns = ['email', 'nombre', 'primer_apellido']
        missing_columns = [col for col in required_columns if col not in column_indices]
        if missing_columns:
            return jsonify({
                'error': f'Faltan columnas requeridas: {", ".join(missing_columns)}',
                'hint': 'Las columnas requeridas son: email, nombre, primer_apellido'
            }), 400
        
        # Función para generar contraseña aleatoria
        def generate_password(length=10):
            alphabet = string.ascii_letters + string.digits
            # Asegurar al menos una mayúscula, una minúscula y un número
            password = [
                secrets.choice(string.ascii_uppercase),
                secrets.choice(string.ascii_lowercase),
                secrets.choice(string.digits)
            ]
            password += [secrets.choice(alphabet) for _ in range(length - 3)]
            secrets.SystemRandom().shuffle(password)
            return ''.join(password)
        
        # Procesar filas
        results = {
            'created': [],
            'errors': [],
            'skipped': [],
            'total_processed': 0
        }
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            results['total_processed'] += 1
            
            # Obtener valores de la fila
            def get_cell_value(field):
                if field not in column_indices:
                    return None
                idx = column_indices[field]
                if idx < len(row):
                    value = row[idx].value
                    return str(value).strip() if value is not None else None
                return None
            
            email = get_cell_value('email')
            nombre = get_cell_value('nombre')
            primer_apellido = get_cell_value('primer_apellido')
            segundo_apellido = get_cell_value('segundo_apellido')
            genero = get_cell_value('genero')
            curp = get_cell_value('curp')
            telefono = get_cell_value('telefono')
            password = get_cell_value('password')
            
            # Validar campos requeridos
            if not email or not nombre or not primer_apellido:
                results['errors'].append({
                    'row': row_idx,
                    'email': email or '(vacío)',
                    'error': 'Campos requeridos vacíos (email, nombre, primer_apellido)'
                })
                continue
            
            # Normalizar email
            email = email.lower().strip()
            
            # Validar formato de email
            if not validate_email(email):
                results['errors'].append({
                    'row': row_idx,
                    'email': email,
                    'error': 'Formato de email inválido'
                })
                continue
            
            # Verificar si el email ya existe
            if User.query.filter_by(email=email).first():
                results['skipped'].append({
                    'row': row_idx,
                    'email': email,
                    'reason': 'Email ya registrado'
                })
                continue
            
            # Verificar CURP si se proporciona
            if curp:
                curp = curp.upper().strip()
                if len(curp) != 18:
                    results['errors'].append({
                        'row': row_idx,
                        'email': email,
                        'error': f'CURP inválido: debe tener 18 caracteres'
                    })
                    continue
                if User.query.filter_by(curp=curp).first():
                    results['skipped'].append({
                        'row': row_idx,
                        'email': email,
                        'reason': f'CURP {curp} ya registrado'
                    })
                    continue
            
            # Normalizar género
            if genero:
                genero = genero.upper()[0] if genero else None
                if genero not in ['M', 'F', 'O']:
                    genero = None
            
            # Generar contraseña si no se proporciona
            generated_password = None
            if not password:
                generated_password = generate_password()
                password = generated_password
            else:
                # Validar contraseña proporcionada
                is_valid, error_msg = validate_password(password)
                if not is_valid:
                    results['errors'].append({
                        'row': row_idx,
                        'email': email,
                        'error': f'Contraseña inválida: {error_msg}'
                    })
                    continue
            
            # Generar username único
            base_username = email.split('@')[0]
            username = base_username
            counter = 1
            while User.query.filter_by(username=username).first():
                username = f"{base_username}{counter}"
                counter += 1
            
            # Crear usuario
            try:
                new_user = User(
                    id=str(uuid.uuid4()),
                    email=email,
                    username=username,
                    name=nombre,
                    first_surname=primer_apellido,
                    second_surname=segundo_apellido or None,
                    gender=genero,
                    curp=curp or None,
                    phone=telefono or None,
                    role='candidato',
                    is_active=True,
                    is_verified=False
                )
                new_user.set_password(password)
                db.session.add(new_user)
                
                results['created'].append({
                    'row': row_idx,
                    'email': email,
                    'name': f"{nombre} {primer_apellido}",
                    'username': username,
                    'password': generated_password  # Solo si fue generada automáticamente
                })
            except Exception as e:
                results['errors'].append({
                    'row': row_idx,
                    'email': email,
                    'error': str(e)
                })
                continue
        
        # Commit de todos los usuarios creados
        if results['created']:
            db.session.commit()
        
        return jsonify({
            'message': f'Proceso completado: {len(results["created"])} usuarios creados',
            'summary': {
                'total_processed': results['total_processed'],
                'created': len(results['created']),
                'errors': len(results['errors']),
                'skipped': len(results['skipped'])
            },
            'details': results
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/candidates/bulk-upload/template', methods=['GET'])
@jwt_required()
@management_required
def download_bulk_upload_template():
    """
    Descargar plantilla Excel para carga masiva de candidatos
    """
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import send_file
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidatos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        required_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        optional_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            ('email', 'Requerido'),
            ('nombre', 'Requerido'),
            ('primer_apellido', 'Requerido'),
            ('segundo_apellido', 'Opcional'),
            ('genero', 'Opcional (M, F, O)'),
            ('curp', 'Opcional (18 caracteres)'),
            ('telefono', 'Opcional'),
            ('password', 'Opcional (se genera automáticamente si está vacío)')
        ]
        
        for col, (header, _) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Segunda fila con descripción
        for col, (_, description) in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=description)
            if 'Requerido' in description:
                cell.fill = required_fill
            else:
                cell.fill = optional_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        # Ejemplo de datos
        example_data = [
            ('candidato1@email.com', 'Juan', 'García', 'López', 'M', 'GALJ900101HDFRPR01', '5512345678', ''),
            ('candidato2@email.com', 'María', 'Pérez', 'Sánchez', 'F', '', '5598765432', 'MiPassword123'),
        ]
        
        for row_idx, data in enumerate(example_data, start=3):
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Ajustar ancho de columnas
        column_widths = [30, 15, 18, 18, 15, 25, 15, 35]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Hoja de instrucciones
        ws_instructions = wb.create_sheet("Instrucciones")
        instructions = [
            "INSTRUCCIONES PARA CARGA MASIVA DE CANDIDATOS",
            "",
            "1. Use la hoja 'Candidatos' para ingresar los datos",
            "2. No modifique los encabezados de las columnas",
            "3. Elimine las filas de ejemplo antes de agregar sus datos",
            "",
            "CAMPOS REQUERIDOS:",
            "- email: Correo electrónico único del candidato",
            "- nombre: Nombre(s) del candidato",
            "- primer_apellido: Apellido paterno",
            "",
            "CAMPOS OPCIONALES:",
            "- segundo_apellido: Apellido materno",
            "- genero: M (Masculino), F (Femenino), O (Otro)",
            "- curp: CURP del candidato (18 caracteres)",
            "- telefono: Número de teléfono",
            "- password: Contraseña (si está vacío, se genera automáticamente)",
            "",
            "REQUISITOS DE CONTRASEÑA:",
            "- Mínimo 8 caracteres",
            "- Al menos una mayúscula",
            "- Al menos una minúscula",
            "- Al menos un número",
            "",
            "NOTAS:",
            "- Los emails duplicados serán omitidos",
            "- Los CURP duplicados serán omitidos",
            "- Se genera un reporte con los resultados de la carga"
        ]
        
        for row_idx, text in enumerate(instructions, start=1):
            cell = ws_instructions.cell(row=row_idx, column=1, value=text)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif text.endswith(':'):
                cell.font = Font(bold=True)
        
        ws_instructions.column_dimensions['A'].width = 60
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='plantilla_candidatos.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
